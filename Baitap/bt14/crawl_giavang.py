# Bài tập 14: Lấy giá vàng từ trang giavang.org và lưu ra file Excel

from datetime import datetime

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

url = "https://giavang.org"
headers = {"User-Agent": "Mozilla/5.0"}

print("Dang tai trang...")
res = requests.get(url, headers=headers)
html = res.text
print("Tai xong!")

soup = BeautifulSoup(html, "html.parser")
tables = soup.find_all("table")
print("So bang tim duoc:", len(tables))

wb = Workbook()
wb.remove(wb.active)


# Lấy tiêu đề h2/h3 gần nhất phía trên bảng làm tên sheet
def lay_tieu_de(table):
    prev = table.find_all_previous(["h1", "h2", "h3"])
    for p in prev:
        t = p.get_text(" ", strip=True)
        if t:
            return t
    return "Bang gia"


# Excel không cho một số ký tự trong tên sheet, tối đa 31 ký tự
def sheet_name_hop_le(name, da_dung):
    for ch in "\\/*?:[]":
        name = name.replace(ch, "")
    name = name[:28]
    if name == "":
        name = "Sheet"
    goc = name
    k = 1
    while name in da_dung:
        k = k + 1
        name = goc + "_" + str(k)
    da_dung.append(name)
    return name


ten_da_dung = []
for i in range(len(tables)):
    table = tables[i]
    tieu_de = lay_tieu_de(table)
    ten_sheet = sheet_name_hop_le(tieu_de, ten_da_dung)
    sheet = wb.create_sheet(ten_sheet)

    rows = table.find_all("tr")

    # Tạo lưới 2 chiều để xử lý ô gộp (rowspan/colspan)
    grid = []
    for _ in range(len(rows)):
        grid.append([])

    for r in range(len(rows)):
        tr = rows[r]
        cells = tr.find_all(["th", "td"])
        col = 0
        for cell in cells:
            # Bỏ qua ô đã bị ô khác gộp xuống
            while col < len(grid[r]) and grid[r][col] is not None:
                col = col + 1

            text = cell.get_text(" ", strip=True)
            text = " ".join(text.split())

            rowspan = cell.get("rowspan")
            if rowspan is None:
                rowspan = 1
            else:
                rowspan = int(rowspan)

            colspan = cell.get("colspan")
            if colspan is None:
                colspan = 1
            else:
                colspan = int(colspan)

            # Điền giá trị vào tất cả ô mà ô này chiếm (xử lý ô gộp)
            for dr in range(rowspan):
                rr = r + dr
                if rr >= len(grid):
                    continue
                while len(grid[rr]) < col + colspan:
                    grid[rr].append(None)
                for dc in range(colspan):
                    grid[rr][col + dc] = text

            col = col + colspan

    sheet.append(["Nguon: " + url])
    sheet.append(["Thoi diem crawl: " + datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
    sheet.append([tieu_de])
    sheet.append([])
    dong_header = sheet.max_row + 1

    so_cot = 0
    for row in grid:
        new_row = []
        for c in row:
            if c is None:
                new_row.append("")
            else:
                new_row.append(c)
        sheet.append(new_row)
        if len(new_row) > so_cot:
            so_cot = len(new_row)

    # Tô màu dòng header của bảng
    mau_vang = PatternFill("solid", fgColor="FFE699")
    font_dam = Font(bold=True)
    canh_giua = Alignment(horizontal="center", wrap_text=True)
    for c in range(1, so_cot + 1):
        o = sheet.cell(row=dong_header, column=c)
        o.font = font_dam
        o.fill = mau_vang
        o.alignment = canh_giua

    for c in range(1, so_cot + 1):
        sheet.column_dimensions[get_column_letter(c)].width = 22
    sheet.freeze_panes = sheet.cell(row=dong_header + 1, column=1)

    print("Da ghi bang", i + 1, "voi", len(grid), "dong x", so_cot, "cot")

wb.save("gia_vang.xlsx")
print("Da luu file gia_vang.xlsx")
